import openpyxl
from openpyxl.drawing.image import Image
from pathlib import Path

# Function to add images to Excel
def add_images_to_excel(image_folder, excel_file):
    # Create a new workbook or load existing
    if Path(excel_file).exists():
        workbook = openpyxl.load_workbook(excel_file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Iterate over PNG images in the specified folder
    for image_path in Path(image_folder).glob('*.png'):
        # Load the image
        img = Image(image_path)

        # Add a new row at the end
        new_row = sheet.max_row + 1

        # Adjust the size of the row and column to better fit the image
        # These values can be adjusted as needed
        sheet.row_dimensions[new_row].height = img.height // 6  # Adjust the row height
        sheet.column_dimensions['A'].width = img.width // 6     # Adjust the column width

        # Add the image to the worksheet, positioned at the top-left of the new row
        sheet.add_image(img, f'A{new_row}')

   
    # Save the workbook
    workbook.save(excel_file)

# Example usage

image_folder = '/Users/charmibhatt/Library/CloudStorage/OneDrive-TheUniversityofWesternOntario/UWO_onedrive/Local_GitHub/DIBs/plotting_CH_plus_lines_linewidth_check/continuum_subtracted'
excel_file = '/Users/charmibhatt/Library/CloudStorage/OneDrive-TheUniversityofWesternOntario/UWO_onedrive/Local_GitHub/DIBs/plotting_CH_plus_lines_linewidth_check/continuum_subtracted/Dtermining_continuum_for_CH_plus.xlsx'

add_images_to_excel(image_folder, excel_file)
